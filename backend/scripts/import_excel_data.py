#!/usr/bin/env python3
"""
Excel数据导入脚本
导入导师评价数据到数据库
"""

import sys
import os
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import asyncio

# 添加项目根目录到Python路径
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select
from app.core.config import settings
from app.models.professor import Professor
from app.models.professor_evaluation import ProfessorEvaluation

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 数据库引擎
engine = create_async_engine(str(settings.DATABASE_URL))
AsyncSessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


class ExcelDataImporter:
    """Excel数据导入器"""

    def __init__(self, file_path: str, limit: Optional[int] = None):
        self.file_path = file_path
        self.limit = limit
        self.wb = None
        self.ws = None

    def load(self) -> bool:
        """加载Excel文件"""
        try:
            self.wb = load_workbook(filename=self.file_path, data_only=True)
            self.ws = self.wb.active
            logger.info(f"成功加载Excel文件: {self.file_path}")
            logger.info(f"工作表: {self.ws.title}, 行数: {self.ws.max_row}")
            return True
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            return False

    def parse_row(self, row_idx: int) -> Optional[Dict[str, Any]]:
        """解析单行数据"""
        try:
            # 根据已知列索引解析
            university = self.ws.cell(row=row_idx, column=1).value
            department = self.ws.cell(row=row_idx, column=2).value
            name = self.ws.cell(row=row_idx, column=3).value
            score = self.ws.cell(row=row_idx, column=4).value
            comments = self.ws.cell(row=row_idx, column=5).value

            # 检查必要字段
            if not name or not university:
                return None

            # 清理数据
            name = str(name).strip()
            university = str(university).strip()
            department = str(department).strip() if department else None

            # 解析分数
            personality_score = None
            if score is not None:
                try:
                    score_val = float(score)
                    if 1 <= score_val <= 5:
                        personality_score = score_val
                except (ValueError, TypeError):
                    pass

            # 解析评价
            student_comments = str(comments).strip() if comments else None

            return {
                "professor": {
                    "name": name,
                    "university": university,
                    "department": department,
                    "research_fields": [],  # 暂时为空
                },
                "evaluation": {
                    "personality_score": personality_score,
                    "student_comments": student_comments,
                    "source": "excel",
                }
            }
        except Exception as e:
            logger.error(f"解析第 {row_idx} 行失败: {e}")
            return None

    async def import_to_database(self) -> Dict[str, Any]:
        """导入数据到数据库"""
        if not self.load():
            return {"success": False, "error": "无法加载Excel文件"}

        total_rows = self.ws.max_row
        start_row = 2  # 跳过表头
        end_row = total_rows + 1

        if self.limit:
            end_row = min(start_row + self.limit, total_rows + 1)

        logger.info(f"准备导入 {end_row - start_row} 行数据")

        imported_professors = 0
        imported_evaluations = 0
        skipped = 0

        async with AsyncSessionLocal() as session:
            try:
                for row_idx in range(start_row, end_row):
                    row_data = self.parse_row(row_idx)
                    if not row_data:
                        skipped += 1
                        continue

                    prof_data = row_data["professor"]
                    eval_data = row_data["evaluation"]

                    # 检查教授是否已存在
                    query = select(Professor).where(
                        Professor.name == prof_data["name"],
                        Professor.university == prof_data["university"]
                    )
                    result = await session.execute(query)
                    existing_prof = result.scalar_one_or_none()

                    if existing_prof:
                        professor = existing_prof
                    else:
                        # 创建新教授
                        professor = Professor(
                            name=prof_data["name"],
                            university=prof_data["university"],
                            department=prof_data["department"],
                            research_fields=prof_data["research_fields"]
                        )
                        session.add(professor)
                        await session.flush()  # 获取ID
                        imported_professors += 1

                    # 创建评价（如果存在）
                    if eval_data["personality_score"] or eval_data["student_comments"]:
                        evaluation = ProfessorEvaluation(
                            professor_id=professor.id,
                            source=eval_data["source"],
                            personality_score=eval_data["personality_score"],
                            student_comments=eval_data["student_comments"]
                        )
                        session.add(evaluation)
                        imported_evaluations += 1

                    # 每100行提交一次
                    if row_idx % 100 == 0:
                        await session.commit()
                        logger.info(f"已处理 {row_idx - start_row + 1} 行...")

                # 提交剩余数据
                await session.commit()

                logger.info(f"数据导入完成")
                logger.info(f"导入教授: {imported_professors}")
                logger.info(f"导入评价: {imported_evaluations}")
                logger.info(f"跳过行: {skipped}")

                return {
                    "success": True,
                    "imported_professors": imported_professors,
                    "imported_evaluations": imported_evaluations,
                    "skipped_rows": skipped,
                    "total_rows_processed": end_row - start_row,
                }

            except Exception as e:
                await session.rollback()
                logger.error(f"导入数据失败: {e}")
                return {"success": False, "error": str(e)}


async def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description="Excel数据导入工具")
    parser.add_argument("--file", required=True, help="Excel文件路径")
    parser.add_argument("--limit", type=int, help="限制导入的行数（用于测试）")
    parser.add_argument("--init-db", action="store_true", help="先初始化数据库")

    args = parser.parse_args()

    # 检查文件是否存在
    if not os.path.exists(args.file):
        print(f"错误: 文件不存在: {args.file}")
        sys.exit(1)

    # 如果需要，初始化数据库
    if args.init_db:
        from scripts.init_database import init_database
        logger.info("初始化数据库...")
        success = await init_database()
        if not success:
            print("数据库初始化失败")
            sys.exit(1)

    # 导入数据
    importer = ExcelDataImporter(args.file, args.limit)
    result = await importer.import_to_database()

    if result["success"]:
        print(f"\n=== 导入成功 ===")
        print(f"导入教授: {result['imported_professors']}")
        print(f"导入评价: {result['imported_evaluations']}")
        print(f"处理行数: {result['total_rows_processed']}")
        sys.exit(0)
    else:
        print(f"\n=== 导入失败 ===")
        print(f"错误: {result.get('error', '未知错误')}")
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())