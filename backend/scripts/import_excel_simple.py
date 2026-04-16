#!/usr/bin/env python3
"""
简化版Excel数据导入脚本
使用openpyxl替代pandas，避免依赖问题
"""

import sys
import os
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import re

# 添加项目根目录到Python路径
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SimpleExcelParser:
    """简化版Excel解析器（使用openpyxl）"""

    # 支持的列名映射
    COLUMN_MAPPINGS = {
        # 导师基本信息
        "name": ["姓名", "导师姓名", "教授姓名", "老师姓名", "name", "professor_name"],
        "university": ["学校", "大学", "高校", "university", "school"],
        "department": ["院系", "学院", "部门", "department", "college"],
        "title": ["职称", "职位", "title", "position"],
        "research_fields": ["研究方向", "研究领域", "研究兴趣", "research_fields", "research_areas"],

        # 评价信息
        "personality_score": ["人品得分", "人品分数", "人品评价", "personality_score", "character_score"],
        "group_atmosphere": ["课题组氛围", "实验室氛围", "团队氛围", "group_atmosphere", "lab_atmosphere"],
        "student_comments": ["学生评价", "学生反馈", "学生评论", "student_comments", "student_feedback"],
        "evaluation_date": ["评价日期", "日期", "时间", "evaluation_date", "date"],

        # 数据来源
        "source": ["来源", "数据来源", "source", "data_source"],
    }

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self.headers = []
        self.column_mapping = {}

    def load(self) -> bool:
        """加载Excel文件"""
        try:
            self.wb = load_workbook(filename=self.file_path, data_only=True)
            if self.sheet_name:
                self.ws = self.wb[self.sheet_name]
            else:
                self.ws = self.wb.active

            # 读取表头（第一行）
            self.headers = []
            for cell in self.ws[1]:
                self.headers.append(cell.value)

            logger.info(f"成功加载Excel文件: {self.file_path}")
            logger.info(f"工作表: {self.ws.title}, 行数: {self.ws.max_row}, 列数: {self.ws.max_column}")
            return True
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            return False

    def detect_columns(self) -> Dict[str, str]:
        """检测列名映射"""
        if not self.headers:
            return {}

        column_mapping = {}

        for standard_name, possible_names in self.COLUMN_MAPPINGS.items():
            for idx, header in enumerate(self.headers):
                if header is None:
                    continue

                header_str = str(header).strip().lower()
                for possible_name in possible_names:
                    possible_name_lower = possible_name.lower().strip()
                    if header_str == possible_name_lower:
                        column_mapping[standard_name] = idx  # 存储列索引
                        break
                if standard_name in column_mapping:
                    break

        self.column_mapping = column_mapping
        logger.info(f"检测到的列名映射: {column_mapping}")
        return column_mapping

    def get_cell_value(self, row, col_idx):
        """获取单元格值"""
        try:
            cell = self.ws.cell(row=row, column=col_idx + 1)  # openpyxl列从1开始
            return cell.value
        except:
            return None

    def parse_row(self, row_idx: int) -> Dict[str, Any]:
        """解析单行数据"""
        data = {
            "professor_info": {},
            "evaluation_info": {},
            "source_info": {},
        }

        # 解析导师基本信息
        if "name" in self.column_mapping:
            name = self.get_cell_value(row_idx, self.column_mapping["name"])
            if name:
                data["professor_info"]["name"] = str(name).strip()

        if "university" in self.column_mapping:
            university = self.get_cell_value(row_idx, self.column_mapping["university"])
            if university:
                data["professor_info"]["university"] = str(university).strip()

        if "department" in self.column_mapping:
            dept = self.get_cell_value(row_idx, self.column_mapping["department"])
            if dept:
                data["professor_info"]["department"] = str(dept).strip()

        if "title" in self.column_mapping:
            title = self.get_cell_value(row_idx, self.column_mapping["title"])
            if title:
                data["professor_info"]["title"] = str(title).strip()

        # 解析研究方向
        if "research_fields" in self.column_mapping:
            research = self.get_cell_value(row_idx, self.column_mapping["research_fields"])
            if research:
                research_str = str(research).strip()
                research_fields = re.split(r'[,;，；\n]', research_str)
                research_fields = [field.strip() for field in research_fields if field.strip()]
                data["professor_info"]["research_fields"] = research_fields

        # 解析评价信息
        if "personality_score" in self.column_mapping:
            score = self.get_cell_value(row_idx, self.column_mapping["personality_score"])
            if score is not None:
                try:
                    score_val = float(score)
                    if 1 <= score_val <= 5:
                        data["evaluation_info"]["personality_score"] = score_val
                except (ValueError, TypeError):
                    pass

        if "group_atmosphere" in self.column_mapping:
            atmosphere = self.get_cell_value(row_idx, self.column_mapping["group_atmosphere"])
            if atmosphere:
                data["evaluation_info"]["group_atmosphere"] = str(atmosphere).strip()

        if "student_comments" in self.column_mapping:
            comments = self.get_cell_value(row_idx, self.column_mapping["student_comments"])
            if comments:
                data["evaluation_info"]["student_comments"] = str(comments).strip()

        if "evaluation_date" in self.column_mapping:
            date_val = self.get_cell_value(row_idx, self.column_mapping["evaluation_date"])
            if date_val:
                try:
                    if isinstance(date_val, datetime):
                        data["evaluation_info"]["evaluation_date"] = date_val.date()
                    elif isinstance(date_val, str):
                        # 尝试解析日期字符串
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"]:
                            try:
                                parsed_date = datetime.strptime(date_val, fmt).date()
                                data["evaluation_info"]["evaluation_date"] = parsed_date
                                break
                            except ValueError:
                                continue
                except Exception as e:
                    logger.warning(f"解析日期失败: {date_val}, 错误: {e}")

        # 解析数据来源
        if "source" in self.column_mapping:
            source = self.get_cell_value(row_idx, self.column_mapping["source"])
            if source:
                data["source_info"]["source"] = str(source).strip()
        else:
            data["source_info"]["source"] = "excel"

        return data

    def parse_all(self) -> List[Dict[str, Any]]:
        """解析所有数据"""
        if not self.ws:
            logger.error("请先加载Excel文件")
            return []

        # 检测列名映射
        self.detect_columns()

        # 检查必要的列
        required_columns = ["name", "university"]
        missing_columns = [col for col in required_columns if col not in self.column_mapping]

        if missing_columns:
            logger.warning(f"缺少必要的列: {missing_columns}，可能无法正确解析所有数据")

        # 解析每一行（从第2行开始）
        parsed_data = []
        for row_idx in range(2, self.ws.max_row + 1):
            try:
                row_data = self.parse_row(row_idx)

                # 只保留有基本信息的行
                if (row_data["professor_info"].get("name") and
                    row_data["professor_info"].get("university")):
                    parsed_data.append(row_data)
                else:
                    logger.warning(f"跳过第 {row_idx} 行: 缺少必要信息")
            except Exception as e:
                logger.error(f"解析第 {row_idx} 行失败: {e}")

        logger.info(f"成功解析 {len(parsed_data)} 条数据")
        return parsed_data

    def get_summary(self) -> Dict[str, Any]:
        """获取数据摘要"""
        if not self.ws:
            return {}

        # 统计非空值
        data_types = {}
        for col_idx, header in enumerate(self.headers):
            if header is None:
                continue

            values = []
            non_null_count = 0
            for row_idx in range(2, min(100, self.ws.max_row + 1)):  # 只检查前100行
                value = self.get_cell_value(row_idx, col_idx)
                if value is not None:
                    non_null_count += 1
                    values.append(str(value))

            total_count = min(99, self.ws.max_row - 1)
            null_percentage = ((total_count - non_null_count) / total_count * 100) if total_count > 0 else 0

            data_types[str(header)] = {
                "non_null_count": non_null_count,
                "null_percentage": round(null_percentage, 2),
                "sample_values": values[:3],
            }

        summary = {
            "file_path": self.file_path,
            "total_rows": self.ws.max_row - 1,  # 减去表头
            "columns": [str(h) for h in self.headers if h is not None],
            "column_mapping": {k: self.headers[v] for k, v in self.column_mapping.items()},
            "missing_columns": [],
            "data_types": data_types,
        }

        # 检查缺少的列
        required_columns = ["name", "university"]
        summary["missing_columns"] = [col for col in required_columns if col not in self.column_mapping]

        return summary


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description="简化版Excel数据导入工具")
    parser.add_argument("--file", required=True, help="Excel文件路径")
    parser.add_argument("--sheet", help="工作表名称")
    parser.add_argument("--summary", action="store_true", help="只显示数据摘要，不导入")
    parser.add_argument("--output", help="输出JSON文件路径")

    args = parser.parse_args()

    # 创建解析器
    parser = SimpleExcelParser(args.file, args.sheet)

    if not parser.load():
        print(f"无法加载Excel文件: {args.file}")
        sys.exit(1)

    # 显示摘要
    summary = parser.get_summary()
    print("\n=== 数据摘要 ===")
    print(f"文件: {summary['file_path']}")
    print(f"总行数: {summary['total_rows']}")
    print(f"列: {', '.join(summary['columns'])}")
    print(f"列映射: {json.dumps(summary['column_mapping'], ensure_ascii=False, indent=2)}")

    if summary['missing_columns']:
        print(f"警告: 缺少必要列: {summary['missing_columns']}")

    if args.summary:
        # 只显示摘要，不解析
        sys.exit(0)

    # 解析数据
    print("\n=== 开始解析数据 ===")
    data = parser.parse_all()

    if not data:
        print("未解析到有效数据")
        sys.exit(1)

    print(f"成功解析 {len(data)} 条数据")

    # 显示前几条数据
    print("\n=== 前3条数据示例 ===")
    for i, item in enumerate(data[:3]):
        print(f"\n数据 {i+1}:")
        print(f"  姓名: {item['professor_info'].get('name')}")
        print(f"  学校: {item['professor_info'].get('university')}")
        print(f"  院系: {item['professor_info'].get('department', 'N/A')}")
        print(f"  研究方向: {item['professor_info'].get('research_fields', [])}")
        if 'personality_score' in item['evaluation_info']:
            print(f"  人品得分: {item['evaluation_info']['personality_score']}")

    # 输出到JSON文件
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        print(f"\n数据已保存到: {args.output}")

    print("\n=== 解析完成 ===")

    # 提供导入建议
    print("\n下一步: 运行导入脚本将数据导入数据库")
    print("命令示例: python import_to_db.py --data parsed_data.json")


if __name__ == "__main__":
    main()