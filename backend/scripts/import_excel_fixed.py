#!/usr/bin/env python3
"""
针对特定Excel格式的数据导入脚本
Excel格式：
1. 学校 (university)
2. 学院 (department)
3. 姓名 (name)
4. 分数 (personality_score, 1-5分)
5. 评价 (student_comments)
"""

import sys
import json
import logging
import re
import asyncio
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# 添加项目根目录到Python路径
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FixedExcelParser:
    """针对特定Excel格式的解析器"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.ws = None

    def load(self) -> bool:
        """加载Excel文件"""
        try:
            self.wb = load_workbook(filename=self.file_path, data_only=True)
            self.ws = self.wb.active
            logger.info(f"成功加载Excel文件: {self.file_path}")
            logger.info(f"工作表: {self.ws.title}, 行数: {self.ws.max_row}, 列数: {self.ws.max_column}")
            return True
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            return False

    def parse_row(self, row_idx: int) -> Optional[Dict[str, Any]]:
        """解析单行数据"""
        try:
            # 根据已知列索引解析
            # 列索引从0开始（openpyxl列从1开始）
            university = self.ws.cell(row=row_idx, column=1).value
            department = self.ws.cell(row=row_idx, column=2).value
            name = self.ws.cell(row=row_idx, column=3).value
            score = self.ws.cell(row=row_idx, column=4).value
            comments = self.ws.cell(row=row_idx, column=5).value

            # 检查必要字段
            if not name or not university:
                return None

            # 清理数据
            name = str(name).strip() if name else ""
            university = str(university).strip() if university else ""
            department = str(department).strip() if department else None

            # 解析分数
            personality_score = None
            if score is not None:
                try:
                    score_val = float(score)
                    if 1 <= score_val <= 5:
                        personality_score = score_val
                except (ValueError, TypeError):
                    pass

            # 解析评价
            student_comments = str(comments).strip() if comments else None

            # 从评价中提取研究方向（如果有相关信息）
            research_fields = self.extract_research_fields(comments)

            return {
                "professor_info": {
                    "name": name,
                    "university": university,
                    "department": department,
                    "research_fields": research_fields,
                },
                "evaluation_info": {
                    "personality_score": personality_score,
                    "student_comments": student_comments,
                    "source": "excel",
                }
            }
        except Exception as e:
            logger.error(f"解析第 {row_idx} 行失败: {e}")
            return None

    def extract_research_fields(self, comments: Optional[str]) -> List[str]:
        """从评价中提取研究方向"""
        if not comments:
            return []

        # 常见研究方向关键词
        research_keywords = [
            "机器学习", "人工智能", "深度学习", "计算机视觉", "自然语言处理",
            "数据挖掘", "大数据", "云计算", "物联网", "区块链",
            "网络安全", "软件工程", "数据库", "操作系统", "计算机网络",
            "算法", "数据结构", "编程语言", "编译原理", "计算机图形学",
            "机器人", "自动化", "控制理论", "信号处理", "通信工程",
            "电子工程", "微电子", "集成电路", "光电工程", "材料科学",
            "生物医学", "化学工程", "机械工程", "土木工程", "航空航天",
            "力学", "物理", "数学", "统计学", "经济学",
            "管理学", "心理学", "教育学", "法学", "文学"
        ]

        found_fields = []
        comments_lower = comments.lower()

        # 简单关键词匹配
        for keyword in research_keywords:
            if keyword in comments:
                found_fields.append(keyword)

        # 如果没找到，尝试从评价文本中提取
        if not found_fields:
            # 这里可以添加更复杂的提取逻辑
            pass

        return found_fields[:5]  # 最多返回5个研究方向

    def parse_all(self, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        """解析所有数据"""
        if not self.ws:
            logger.error("请先加载Excel文件")
            return []

        parsed_data = []
        total_rows = self.ws.max_row

        # 从第2行开始（跳过表头）
        start_row = 2
        end_row = total_rows + 1

        if limit:
            end_row = min(start_row + limit, total_rows + 1)

        logger.info(f"解析行范围: {start_row} 到 {end_row-1}")

        for row_idx in range(start_row, end_row):
            row_data = self.parse_row(row_idx)
            if row_data:
                parsed_data.append(row_data)

            # 进度提示
            if row_idx % 1000 == 0:
                logger.info(f"已解析 {row_idx - start_row + 1} 行...")

        logger.info(f"成功解析 {len(parsed_data)} 条有效数据（共 {end_row - start_row} 行）")
        return parsed_data


def save_to_json(data: List[Dict[str, Any]], output_path: str):
    """保存数据到JSON文件"""
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"数据已保存到: {output_path}")
        return True
    except Exception as e:
        logger.error(f"保存JSON文件失败: {e}")
        return False


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description="针对特定Excel格式的数据导入工具")
    parser.add_argument("--file", required=True, help="Excel文件路径")
    parser.add_argument("--limit", type=int, help="限制解析的行数（用于测试）")
    parser.add_argument("--output", default="parsed_data.json", help="输出JSON文件路径")
    parser.add_argument("--summary", action="store_true", help="只显示摘要")

    args = parser.parse_args()

    # 创建解析器
    parser = FixedExcelParser(args.file)

    if not parser.load():
        print(f"无法加载Excel文件: {args.file}")
        sys.exit(1)

    # 显示文件信息
    print(f"\n=== 文件信息 ===")
    print(f"文件: {args.file}")
    print(f"工作表: {parser.ws.title}")
    print(f"总行数: {parser.ws.max_row}")
    print(f"总列数: {parser.ws.max_column}")

    # 显示表头
    headers = []
    for col in range(1, min(6, parser.ws.max_column + 1)):
        header = parser.ws.cell(row=1, column=col).value
        headers.append(header)
    print(f"表头: {headers}")

    if args.summary:
        # 显示前几行数据示例
        print(f"\n=== 前5行数据示例 ===")
        for row_idx in range(2, 7):
            if row_idx > parser.ws.max_row:
                break
            data = parser.parse_row(row_idx)
            if data:
                prof = data["professor_info"]
                eval_info = data["evaluation_info"]
                print(f"\n行 {row_idx}:")
                print(f"  姓名: {prof.get('name')}")
                print(f"  学校: {prof.get('university')}")
                print(f"  院系: {prof.get('department')}")
                print(f"  分数: {eval_info.get('personality_score')}")
                if eval_info.get('student_comments'):
                    comments = eval_info['student_comments']
                    preview = comments[:100] + "..." if len(comments) > 100 else comments
                    print(f"  评价预览: {preview}")
        sys.exit(0)

    # 解析数据
    print(f"\n=== 开始解析数据 ===")
    data = parser.parse_all(args.limit)

    if not data:
        print("未解析到有效数据")
        sys.exit(1)

    # 保存到JSON
    if save_to_json(data, args.output):
        print(f"\n=== 解析完成 ===")
        print(f"成功解析 {len(data)} 条数据")
        print(f"数据已保存到: {args.output}")

        # 显示统计信息
        universities = set()
        departments = set()
        scores = []

        for item in data:
            prof = item["professor_info"]
            eval_info = item["evaluation_info"]
            universities.add(prof.get("university"))
            if prof.get("department"):
                departments.add(prof.get("department"))
            if eval_info.get("personality_score"):
                scores.append(eval_info["personality_score"])

        print(f"\n=== 统计信息 ===")
        print(f"学校数量: {len(universities)}")
        print(f"院系数量: {len(departments)}")
        print(f"有效分数数量: {len(scores)}")
        if scores:
            avg_score = sum(scores) / len(scores)
            print(f"平均分数: {avg_score:.2f}")

        print(f"\n=== 前5所学校 ===")
        for i, uni in enumerate(list(universities)[:5]):
            print(f"  {i+1}. {uni}")

    else:
        print("保存数据失败")
        sys.exit(1)


if __name__ == "__main__":
    main()