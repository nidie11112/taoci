#!/usr/bin/env python3
"""
调试Excel文件结构
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

def debug_excel(file_path):
    """调试Excel文件"""
    print(f"打开文件: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    print(f"工作表名称: {ws.title}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

    # 读取表头
    headers = []
    print("\n=== 表头（第一行）===")
    for col_idx, cell in enumerate(ws[1], 1):
        value = cell.value
        headers.append(value)
        print(f"列{col_idx}: {repr(value)} (类型: {type(value)})")

    # 读取前几行数据
    print("\n=== 前5行数据 ===")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        print(f"\n行 {row_idx}:")
        row_data = []
        for col_idx, cell in enumerate(ws[row_idx], 1):
            value = cell.value
            row_data.append(value)
            print(f"  列{col_idx}: {repr(value)}")

        # 显示为字典格式
        row_dict = {}
        for i, header in enumerate(headers):
            if header and i < len(row_data):
                row_dict[header] = row_data[i]

        print(f"  字典格式: {row_dict}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("用法: python debug_excel.py <excel文件路径>")
        sys.exit(1)

    debug_excel(sys.argv[1])